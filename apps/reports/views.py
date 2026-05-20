from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.utils import timezone
from apps.transactions.models import Transaksi, DetailTransaksi
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_bulan_nama(bulan):
    return datetime.date(2000, bulan, 1).strftime('%B')


@login_required
def laporan_keuangan(request):
    bulan = int(request.GET.get('bulan', timezone.now().month))
    tahun = int(request.GET.get('tahun', timezone.now().year))
    is_print = request.GET.get('print', False)
    is_pdf = request.GET.get('pdf', False)

    transaksi = Transaksi.objects.filter(
        status='selesai',
        tanggal_kembali_aktual__year=tahun,
        tanggal_kembali_aktual__month=bulan
    )

    total_omzet = transaksi.aggregate(total=Sum('total_harga'))['total'] or 0
    total_dp = transaksi.aggregate(total=Sum('uang_muka'))['total'] or 0
    total_sisa = transaksi.aggregate(total=Sum('sisa_bayar'))['total'] or 0
    total_transaksi = transaksi.count()

    months = [(i, datetime.date(2000, i, 1).strftime('%B')) for i in range(1, 13)]
    years = list(range(timezone.now().year - 2, timezone.now().year + 2))

    context = {
        'total_omzet': total_omzet,
        'total_dp': total_dp,
        'total_sisa': total_sisa,
        'total_transaksi': total_transaksi,
        'transaksi_list': transaksi.order_by('tanggal_kembali_aktual'),
        'bulan': bulan,
        'tahun': tahun,
        'bulan_nama': get_bulan_nama(bulan),
        'months': months,
        'years': years,
        'is_print': is_print or is_pdf,
        'generated_at': timezone.now(),
    }

    if is_pdf:
        return render(request, 'reports/keuangan_print.html', context)
    return render(request, 'reports/keuangan.html', context)


@login_required
def laporan_barang(request):
    is_print = request.GET.get('print', False)
    is_pdf = request.GET.get('pdf', False)

    detail = DetailTransaksi.objects.select_related('barang', 'barang__kategori').values(
        'barang__id', 'barang__nama', 'barang__kode', 'barang__kategori__nama'
    ).annotate(
        total_disewa=Sum('jumlah'),
        total_transaksi=Count('transaksi', distinct=True),
        total_pendapatan=Sum('subtotal')
    ).order_by('-total_disewa')

    context = {
        'detail_list': detail,
        'is_print': is_print or is_pdf,
        'generated_at': timezone.now(),
    }

    if is_pdf:
        return render(request, 'reports/barang_print.html', context)
    return render(request, 'reports/barang.html', context)


def style_header(cell, bg_color='C9A84C', font_color='FFFFFF'):
    """Helper untuk styling header cell"""
    cell.font = Font(bold=True, color=font_color, size=11)
    cell.fill = PatternFill(start_color=bg_color,
        end_color=bg_color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def style_cell(cell, bold=False, align='left'):
    """Helper untuk styling data cell"""
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


@login_required
def laporan_keuangan(request):
    bulan = int(request.GET.get('bulan', timezone.now().month))
    tahun = int(request.GET.get('tahun', timezone.now().year))
    is_print = request.GET.get('print', False)
    is_pdf = request.GET.get('pdf', False)
    is_excel = request.GET.get('excel', False)

    transaksi = Transaksi.objects.filter(
        status='selesai',
        tanggal_kembali_aktual__year=tahun,
        tanggal_kembali_aktual__month=bulan
    )

    total_omzet = transaksi.aggregate(
        total=Sum('total_harga'))['total'] or 0
    total_dp = transaksi.aggregate(
        total=Sum('uang_muka'))['total'] or 0
    total_sisa = transaksi.aggregate(
        total=Sum('sisa_bayar'))['total'] or 0
    total_transaksi = transaksi.count()

    months = [(i, datetime.date(2000, i, 1).strftime('%B'))
        for i in range(1, 13)]
    years = list(range(timezone.now().year - 2, timezone.now().year + 2))

    context = {
        'total_omzet': total_omzet,
        'total_dp': total_dp,
        'total_sisa': total_sisa,
        'total_transaksi': total_transaksi,
        'transaksi_list': transaksi.order_by('tanggal_kembali_aktual'),
        'bulan': bulan,
        'tahun': tahun,
        'bulan_nama': get_bulan_nama(bulan),
        'months': months,
        'years': years,
        'is_print': is_print or is_pdf,
        'generated_at': timezone.now(),
    }

    if is_excel:
        return export_keuangan_excel(transaksi, bulan, tahun,
            total_omzet, total_dp, total_sisa, total_transaksi)
    if is_pdf:
        return render(request, 'reports/keuangan_print.html', context)
    return render(request, 'reports/keuangan.html', context)


def export_keuangan_excel(transaksi, bulan, tahun,
        total_omzet, total_dp, total_sisa, total_transaksi):
    """Export laporan keuangan ke Excel"""
    bulan_nama = get_bulan_nama(bulan)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Laporan {bulan_nama} {tahun}'

    # ===== HEADER JUDUL =====
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Linda Salon — Manajemen Sewa Perlengkapan Pernikahan'
    ws['A1'].font = Font(bold=True, size=14, color='1A1208')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = f'LAPORAN KEUANGAN — {bulan_nama.upper()} {tahun}'
    ws['A2'].font = Font(bold=True, size=12, color='C9A84C')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:H3')
    ws['A3'] = f'Dicetak: {timezone.now().strftime("%d %B %Y %H:%M")} WIB'
    ws['A3'].font = Font(size=9, color='8A7355')
    ws['A3'].alignment = Alignment(horizontal='center')

    # ===== SUMMARY =====
    ws['A5'] = 'RINGKASAN'
    ws['A5'].font = Font(bold=True, size=11, color='C9A84C')

    summary_data = [
        ('Total Omzet', f'Rp {int(total_omzet):,}'),
        ('Total DP Masuk', f'Rp {int(total_dp):,}'),
        ('Total Sisa Bayar', f'Rp {int(total_sisa):,}'),
        ('Jumlah Transaksi', f'{total_transaksi} transaksi'),
    ]

    for i, (label, value) in enumerate(summary_data):
        row = 6 + i
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(size=10, color='8A7355')
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(bold=True, size=10)

    # ===== HEADER TABEL =====
    header_row = 12
    headers = [
        'No', 'No. Transaksi', 'Pelanggan',
        'Acara', 'Tgl Sewa', 'Tgl Kembali',
        'Total', 'DP', 'Sisa Bayar'
    ]
    ws.merge_cells(f'A{header_row - 1}:I{header_row - 1}')
    ws[f'A{header_row - 1}'] = 'RINCIAN TRANSAKSI'
    ws[f'A{header_row - 1}'].font = Font(bold=True, size=11, color='C9A84C')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        style_header(cell)

    # ===== DATA =====
    for idx, trx in enumerate(transaksi.order_by(
            'tanggal_kembali_aktual'), 1):
        row = header_row + idx
        data = [
            idx,
            trx.no_transaksi,
            trx.pelanggan_nama,
            trx.acara or '-',
            trx.tanggal_sewa.strftime('%d/%m/%Y'),
            trx.tanggal_kembali_aktual.strftime('%d/%m/%Y') if trx.tanggal_kembali_aktual else '-',
            int(trx.total_harga),
            int(trx.uang_muka),
            int(trx.sisa_bayar),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            style_cell(cell,
                align='center' if col in [1, 5, 6] else 'left')
            # Warna alternating row
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color='FAF7F0',
                    end_color='FAF7F0', fill_type='solid')

    # ===== TOTAL ROW =====
    total_row = header_row + total_transaksi + 1
    ws.cell(row=total_row, column=1, value='').border = Border(
        top=Side(style='double'))
    ws.merge_cells(
        f'A{total_row}:F{total_row}')
    total_label = ws.cell(row=total_row,
        column=1, value='TOTAL')
    total_label.font = Font(bold=True, size=11)
    total_label.alignment = Alignment(horizontal='right')
    total_label.fill = PatternFill(start_color='FAF7F0',
        end_color='FAF7F0', fill_type='solid')

    for col, value in enumerate(
            [int(total_omzet), int(total_dp), int(total_sisa)], 7):
        cell = ws.cell(row=total_row, column=col, value=value)
        style_cell(cell, bold=True)
        cell.fill = PatternFill(start_color='FFF8E7',
            end_color='FFF8E7', fill_type='solid')

    # ===== LEBAR KOLOM =====
    col_widths = [5, 20, 25, 20, 12, 12, 18, 18, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ===== ROW HEIGHT =====
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[header_row].height = 20

    # ===== RESPONSE =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="laporan_keuangan_{bulan_nama}_{tahun}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def laporan_barang(request):
    is_print = request.GET.get('print', False)
    is_pdf = request.GET.get('pdf', False)
    is_excel = request.GET.get('excel', False)

    detail = DetailTransaksi.objects.select_related(
        'barang', 'barang__kategori'
    ).values(
        'barang__id', 'barang__nama', 'barang__kode',
        'barang__kategori__nama'
    ).annotate(
        total_disewa=Sum('jumlah'),
        total_transaksi=Count('transaksi', distinct=True),
        total_pendapatan=Sum('subtotal')
    ).order_by('-total_disewa')

    context = {
        'detail_list': detail,
        'is_print': is_print or is_pdf,
        'generated_at': timezone.now(),
    }

    if is_excel:
        return export_barang_excel(detail)
    if is_pdf:
        return render(request, 'reports/barang_print.html', context)
    return render(request, 'reports/barang.html', context)


def export_barang_excel(detail):
    """Export laporan barang ke Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Barang'

    # ===== HEADER JUDUL =====
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Linda Salon — Manajemen Sewa Perlengkapan Pernikahan'
    ws['A1'].font = Font(bold=True, size=14, color='1A1208')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = 'LAPORAN POPULARITAS BARANG'
    ws['A2'].font = Font(bold=True, size=12, color='C9A84C')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:G3')
    ws['A3'] = f'Dicetak: {timezone.now().strftime("%d %B %Y %H:%M")} WIB'
    ws['A3'].font = Font(size=9, color='8A7355')
    ws['A3'].alignment = Alignment(horizontal='center')

    # ===== HEADER TABEL =====
    header_row = 5
    headers = [
        'No', 'Kode', 'Nama Barang',
        'Kategori', 'Total Disewa',
        'Jumlah Transaksi', 'Total Pendapatan'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        style_header(cell)

    # ===== DATA =====
    for idx, item in enumerate(detail, 1):
        row = header_row + idx
        data = [
            idx,
            item['barang__kode'],
            item['barang__nama'],
            item['barang__kategori__nama'] or '-',
            item['total_disewa'],
            item['total_transaksi'],
            int(item['total_pendapatan'] or 0),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            style_cell(cell,
                align='center' if col in [1, 5, 6] else 'left')
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color='FAF7F0',
                    end_color='FAF7F0', fill_type='solid')

    # ===== LEBAR KOLOM =====
    col_widths = [5, 12, 30, 20, 15, 18, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[header_row].height = 20

    # ===== RESPONSE =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="laporan_barang.xlsx"'
    )
    wb.save(response)
    return response
